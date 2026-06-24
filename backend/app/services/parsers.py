import csv
import io
import logging
from typing import Any

logger = logging.getLogger(__name__)

MAX_ROWS = 1000
MAX_ROWS_PER_BATCH = 50
MAX_COLUMNS = 50


def parse_upload(content: bytes, content_type: str | None, filename: str) -> str:
    """
    Extract plain text from uploaded files (PDF, DOCX, TXT, CSV, XLSX).
    
    Args:
        content: File content as bytes
        content_type: MIME type of the file
        filename: Name of the file
        
    Returns:
        Extracted text content
    """
    ct = (content_type or "").lower()
    fn = (filename or "").lower()

    if "pdf" in ct or fn.endswith(".pdf"):
        return _parse_pdf(content)
    if "word" in ct or fn.endswith(".docx") or fn.endswith(".doc"):
        return _parse_docx(content)
    if "csv" in ct or fn.endswith(".csv"):
        return _parse_csv(content)
    if "spreadsheet" in ct or "xlsx" in ct or fn.endswith(".xlsx") or fn.endswith(".xls"):
        return _parse_xlsx(content)
    if ct.startswith("text/") or fn.endswith(".txt") or fn.endswith(".md"):
        return content.decode("utf-8", errors="replace")

    try:
        return _parse_docx(content)
    except Exception:
        return content.decode("utf-8", errors="replace")


def parse_upload_with_metadata(
    content: bytes, content_type: str | None, filename: str
) -> tuple[str, dict[str, Any]]:
    """
    Extract text and return metadata about the parsed file for RAG payload enrichment.
    
    Returns:
        Tuple of (text, metadata_dict) where metadata_dict may contain:
        - file_type: str (e.g. "pdf", "csv", "xlsx", "docx", "txt")
        - file_name: str
        - sheet_name: str (XLSX only)
        - row_range: str (CSV/XLSX only, e.g. "1-50")
    """
    ct = (content_type or "").lower()
    fn = (filename or "").lower()
    file_type = _detect_file_type(ct, fn)
    text = parse_upload(content, content_type, filename)
    meta: dict[str, Any] = {"file_type": file_type, "file_name": filename or ""}
    return text, meta


def _parse_pdf(content: bytes) -> str:
    """
    Extract text from a PDF file.
    
    Args:
        content: PDF file content as bytes
        
    Returns:
        Extracted text content
        
    Raises:
        Exception: If PDF parsing fails
    """
    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(content))
        parts = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if text:
                parts.append(f"--- Page {i + 1} ---\n{text}")
        return "\n\n".join(parts)
    except Exception as exc:
        logger.warning("PDF parse failed: %s", exc)
        return content.decode("utf-8", errors="replace")


def _parse_docx(content: bytes) -> str:
    """
    Extract text from a DOCX file.
    
    Args:
        content: DOCX file content as bytes
        
    Returns:
        Extracted text content
        
    Raises:
        Exception: If DOCX parsing fails
    """
    try:
        from docx import Document

        doc = Document(io.BytesIO(content))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)
    except Exception as exc:
        logger.warning("DOCX parse failed: %s", exc)
        raise


def _detect_file_type(ct: str, fn: str) -> str:
    if "pdf" in ct or fn.endswith(".pdf"):
        return "pdf"
    if "word" in ct or fn.endswith(".docx") or fn.endswith(".doc"):
        return "docx"
    if "csv" in ct or fn.endswith(".csv"):
        return "csv"
    if "spreadsheet" in ct or "xlsx" in ct or fn.endswith(".xlsx") or fn.endswith(".xls"):
        return "xlsx"
    if ct.startswith("text/") or fn.endswith(".txt") or fn.endswith(".md"):
        return "txt"
    return "unknown"


def _parse_csv(content: bytes) -> str:
    """Parse CSV into self-describing key-value text batches with headers."""
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(text.splitlines())
    rows = list(reader)
    if not rows:
        return ""
    headers = [h.strip() for h in rows[0]]
    data_rows = rows[1:]
    return _format_tabular(headers, data_rows)


def _parse_xlsx(content: bytes) -> str:
    """Parse XLSX into self-describing key-value text batches with headers per sheet."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parts = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        data_rows = []
        for row in rows[1:]:
            data_rows.append([str(c) if c is not None else "" for c in row])
        parts.append(_format_tabular(headers, data_rows, sheet_name=ws.title))
    wb.close()
    return "\n\n".join(parts)


def _format_tabular(
    headers: list[str],
    data_rows: list[list[str]],
    sheet_name: str | None = None,
) -> str:
    """Convert tabular data into header-prefixed batches of key-value rows."""
    ncols = min(len(headers), MAX_COLUMNS)
    headers = headers[:ncols]

    truncated = data_rows[:MAX_ROWS]
    if len(data_rows) > MAX_ROWS:
        logger.warning("Table has %d rows, truncating to %d", len(data_rows), MAX_ROWS)

    prefix = f"--- Sheet: {sheet_name} ---\n" if sheet_name else ""
    header_line = f"Headers: {', '.join(headers)}"

    batches = []
    for i in range(0, len(truncated), MAX_ROWS_PER_BATCH):
        batch = truncated[i : i + MAX_ROWS_PER_BATCH]
        lines = [header_line]
        for j, row in enumerate(batch):
            vals = row[:ncols]
            pairs = [
                f"{headers[k]}={vals[k]}"
                for k in range(min(len(vals), ncols))
                if vals[k]
            ]
            if pairs:
                lines.append(f"Row {i + j + 1}: {', '.join(pairs)}")
        batches.append("\n".join(lines))

    return prefix + "\n\n".join(batches)
